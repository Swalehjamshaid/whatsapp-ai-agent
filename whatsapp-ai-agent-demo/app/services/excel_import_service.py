# ==========================================================
# FILE: app/services/excel_import_service.py (v4.1 - SAFE OPTIONAL COLUMN ACCESS)
# PURPOSE: Maps ALL available Excel columns to ALL database fields.
# ==========================================================

import logging
from datetime import datetime, date
from typing import Dict, Any, List, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import DeliveryReport

logger = logging.getLogger(__name__)

# ==========================================================
# CUSTOM EXCEPTIONS
# ==========================================================

class ExcelImportServiceError(Exception):
    """Base exception for Excel import service errors."""
    pass

class VerificationError(Exception):
    """Raised only when critical columns are missing."""
    pass

# ==========================================================
# HELPER FUNCTIONS
# ==========================================================

def _clean_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else None
    return str(value)

def _safe_int(value: Any) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).strip().replace(',', '')))
    except (ValueError, TypeError):
        return None

def _safe_float(value: Any) -> Optional[float]:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).strip().replace(',', ''))
    except (ValueError, TypeError):
        return None

def _parse_date_from_excel(value: Any) -> Optional[date]:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        logger.warning(f"Could not parse date from string: {value}")
    if isinstance(value, (int, float)):
        try:
            from datetime import timedelta
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except Exception:
            pass
    return None

# ==========================================================
# CORE IMPORT FUNCTION (BATCH + FULL MAPPING + SAFE ACCESS)
# ==========================================================

def import_delivery_excel(
    db: Session,
    file_path: str,
    source_filename: str,
    upload_batch_id: str,
    batch_size: int = 1000,
    replace_mode: bool = False
) -> Dict[str, Any]:
    """
    Reads Haier Excel file, maps ALL available columns to the DeliveryReport model,
    and bulk inserts in batches of 1000.
    Missing columns are logged but not fatal (except for dn_no).
    """
    logger.info(f"Starting import for batch: {upload_batch_id}, file: {source_filename}")
    
    metrics = {
        "rows_read": 0,
        "rows_upserted": 0,
        "rows_failed": 0,
        "rows_valid": 0,
        "batch_id": upload_batch_id,
        "errors": []
    }

    # ----------------------------------------------------------
    # STEP 1: LOAD EXCEL WORKBOOK
    # ----------------------------------------------------------
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
    except Exception as e:
        logger.exception(f"Failed to load Excel file: {file_path}")
        raise ExcelImportServiceError(f"Failed to read Excel file: {str(e)}")

    # ----------------------------------------------------------
    # STEP 2: EXTRACT HEADERS (Case‑insensitive)
    # ----------------------------------------------------------
    headers = []
    for cell in ws[1]:
        header_val = _clean_string(cell.value)
        headers.append(header_val.upper() if header_val else None)

    # Define a comprehensive mapping of possible Excel column names to model fields.
    # All columns are optional EXCEPT "DN NO" – if that's missing we raise an error.
    expected_headers = {
        "ORDER TYPE": "order_type",
        "DN NO": "dn_no",                  # REQUIRED
        "DN AMOUNT": "dn_amount",
        "DN QTY": "dn_qty",
        "DN WORK": "dn_work",
        "DIVISION": "division",
        "MATERIAL NO": "material_no",
        "CUSTOMER MODEL": "customer_model",
        "SALES OFFICE": "sales_office",
        "SOLD-TO-PARTY NAME": "customer_name",
        "CUSTOMER CODE": "customer_code",
        "DEALER CODE": "dealer_code",
        "SHIP-TO CITY": "ship_to_city",
        "STORAGE": "storage_location",
        "WAREHOUSE": "warehouse",
        "WAREHOUSE CODE": "warehouse_code",
        "DELIVERY LOCATION": "delivery_location",
        "DN CREATE DATE": "dn_create_date",
        "GOOD ISSUE DATE": "good_issue_date",
        "POD DATE": "pod_date",
        "SALES MANAGER": "sales_manager",
        "REMARKS": "remarks",
    }

    # Build column map, and warn if a column is missing (except for required ones)
    column_map = {}
    missing_columns = []
    for expected_upper, model_field in expected_headers.items():
        found = False
        for i, h in enumerate(headers):
            if h and expected_upper in h:
                column_map[model_field] = i
                found = True
                break
        if not found:
            if model_field == "dn_no":
                # DN NO is mandatory – raise an error if missing
                raise VerificationError(f"Required column '{expected_upper}' (DN NO) not found in Excel header.")
            else:
                # Optional columns – just log a warning and skip mapping
                missing_columns.append(expected_upper)
                logger.warning(f"Optional column '{expected_upper}' not found in Excel, will leave field as NULL.")

    # If many columns are missing, log a summary
    if missing_columns:
        logger.info(f"The following columns were not found and will be left NULL: {', '.join(missing_columns)}")

    # ----------------------------------------------------------
    # STEP 3: PARSE ROWS INTO OBJECTS (SAFE ACCESS FOR OPTIONAL COLUMNS)
    # ----------------------------------------------------------
    records_to_insert = []
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        metrics["rows_read"] += 1
        
        try:
            # Mandatory field
            dn_no = _clean_string(row[column_map["dn_no"]])
            if not dn_no:
                raise ValueError("DN No is empty or missing")

            # Helper to safely get column value
            def _get_cell(model_field):
                idx = column_map.get(model_field)
                return row[idx] if idx is not None else None

            # Optional fields – safe access with _get_cell()
            report = DeliveryReport(
                order_type=_clean_string(_get_cell("order_type")),
                dn_no=dn_no,
                dn_amount=_safe_float(_get_cell("dn_amount")),
                dn_qty=_safe_int(_get_cell("dn_qty")),
                dn_work=_clean_string(_get_cell("dn_work")),
                division=_clean_string(_get_cell("division")),
                material_no=_clean_string(_get_cell("material_no")),
                customer_model=_clean_string(_get_cell("customer_model")),
                sales_office=_clean_string(_get_cell("sales_office")),
                customer_name=_clean_string(_get_cell("customer_name")),
                customer_code=_clean_string(_get_cell("customer_code")),
                dealer_code=_clean_string(_get_cell("dealer_code")),
                ship_to_city=_clean_string(_get_cell("ship_to_city")),
                storage_location=_clean_string(_get_cell("storage_location")),
                warehouse=_clean_string(_get_cell("warehouse")),
                warehouse_code=_clean_string(_get_cell("warehouse_code")),
                delivery_location=_clean_string(_get_cell("delivery_location")),
                dn_create_date=_parse_date_from_excel(_get_cell("dn_create_date")),
                good_issue_date=_parse_date_from_excel(_get_cell("good_issue_date")),
                pod_date=_parse_date_from_excel(_get_cell("pod_date")),
                sales_manager=_clean_string(_get_cell("sales_manager")),
                remarks=_clean_string(_get_cell("remarks")),
                delivery_status="Delivered" if _parse_date_from_excel(_get_cell("pod_date")) else "Pending",
                pgi_status="Completed" if _parse_date_from_excel(_get_cell("good_issue_date")) else "Pending",
                pod_status="Completed" if _parse_date_from_excel(_get_cell("pod_date")) else "Pending",
                pending_flag=False if _parse_date_from_excel(_get_cell("pod_date")) else True,
                source_file=source_filename,
                upload_batch_id=upload_batch_id
            )

            records_to_insert.append(report)
            metrics["rows_valid"] += 1

        except Exception as e:
            logger.warning(f"Row {row_idx} failed validation: {e}")
            errors.append({"row": row_idx, "error": str(e)})
            metrics["rows_failed"] += 1

    # ----------------------------------------------------------
    # STEP 4: BULK INSERT IN BATCHES OF 1000
    # ----------------------------------------------------------
    if records_to_insert:
        try:
            # Prepare dictionaries for bulk insert
            records_dict = [
                {
                    "order_type": r.order_type,
                    "dn_no": r.dn_no,
                    "dn_amount": r.dn_amount,
                    "dn_qty": r.dn_qty,
                    "dn_work": r.dn_work,
                    "division": r.division,
                    "material_no": r.material_no,
                    "customer_model": r.customer_model,
                    "sales_office": r.sales_office,
                    "customer_name": r.customer_name,
                    "customer_code": r.customer_code,
                    "dealer_code": r.dealer_code,
                    "ship_to_city": r.ship_to_city,
                    "storage_location": r.storage_location,
                    "warehouse": r.warehouse,
                    "warehouse_code": r.warehouse_code,
                    "delivery_location": r.delivery_location,
                    "dn_create_date": r.dn_create_date,
                    "good_issue_date": r.good_issue_date,
                    "pod_date": r.pod_date,
                    "sales_manager": r.sales_manager,
                    "remarks": r.remarks,
                    "delivery_status": r.delivery_status,
                    "pgi_status": r.pgi_status,
                    "pod_status": r.pod_status,
                    "pending_flag": r.pending_flag,
                    "source_file": r.source_file,
                    "upload_batch_id": r.upload_batch_id
                }
                for r in records_to_insert
            ]

            total_batches = 0
            for i in range(0, len(records_dict), batch_size):
                batch = records_dict[i:i + batch_size]
                db.bulk_insert_mappings(DeliveryReport, batch)
                db.flush()
                total_batches += len(batch)
                logger.info(f"Inserted batch of {len(batch)} records. Total so far: {total_batches}")

            metrics["rows_upserted"] = len(records_to_insert)
            logger.info(f"✅ Successfully inserted {metrics['rows_upserted']} records into delivery_reports.")

        except Exception as e:
            logger.exception("Failed to bulk insert records.")
            db.rollback()
            raise ExcelImportServiceError(f"Database insert failed: {str(e)}")
    else:
        logger.info("No valid records found to insert.")

    metrics["errors"] = [f"Row {err['row']}: {err['error']}" for err in errors]
    
    logger.info(f"Import completed for batch {upload_batch_id}. "
                f"Read: {metrics['rows_read']}, "
                f"Inserted: {metrics['rows_upserted']}, "
                f"Failed: {metrics['rows_failed']}")

    return metrics
